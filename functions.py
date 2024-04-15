import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def plot_graph(lcoh_pem_pv, lcoh_pem_WindOnshore ,lcoh_pem_WindOffshore, lcoh_pem_nuclear, lcoh_alk_pv, lcoh_alk_WindOnshore, lcoh_alk_WindOffshore, lcoh_alk_nuclear, 
               lcoh_soec_pv,lcoh_soec_WindOnshore, lcoh_soec_WindOffshore, lcoh_soec_nuclear, lcoh_aem_pv, lcoh_aem_WindOnshore, lcoh_aem_WindOffshore, lcoh_aem_nuclear, name):
    # Dados
    eletrolisadores = ['PEM', 'ALK', 'SOEC', 'AEM']
    tecnologias = ['Solar', 'Eólica OnShore', 'Eólica OffShore', 'Nuclear']
    custo_pem = [lcoh_pem_pv, lcoh_pem_WindOnshore, lcoh_pem_WindOffshore, lcoh_pem_nuclear]
    custo_alk = [lcoh_alk_pv, lcoh_alk_WindOnshore, lcoh_alk_WindOffshore, lcoh_alk_nuclear]
    custo_soec = [lcoh_soec_pv, lcoh_soec_WindOnshore, lcoh_soec_WindOffshore, lcoh_soec_nuclear]
    custo_aem = [lcoh_aem_pv, lcoh_aem_WindOnshore, lcoh_aem_WindOffshore, lcoh_aem_nuclear]

    # Cores mais sóbrias
    cores = [
        (46/255, 139/255, 87/255),  # ForestGreen
        (138/255, 43/255, 226/255),  # Blueviolet
        (255/255, 69/255, 0/255),  # Laranja avermelhado
        (255/255, 200/255, 0/255)     # Amarelo
    ]  

    # Configuração do gráfico
    barWidth = 0.2
    fig, ax = plt.subplots(figsize=(10, 6)) 

    # Posição das barras
    r = np.arange(len(tecnologias))
    r1 = r - barWidth
    r2 = r
    r3 = r + barWidth
    r4 = r + 2*barWidth

    # Plotagem das barras
    ax.bar(r1, custo_pem, color=cores[0], width=barWidth, label= eletrolisadores[0]) 
    ax.bar(r2, custo_alk, color=cores[1], width=barWidth, label=eletrolisadores[1]) 
    ax.bar(r3, custo_soec, color=cores[2], width=barWidth, label=eletrolisadores[2]) 
    ax.bar(r4, custo_aem, color=cores[3],width=barWidth, label=eletrolisadores[3]) 

    # Personalização do gráfico
    ax.set_xlabel('Tecnologia', fontweight='bold', fontsize=15) 
    ax.set_ylabel('LCOH ($/kg)', fontweight='bold', fontsize=15) 
    ax.set_xticks(r)
    ax.set_xticklabels(tecnologias)
    ax.legend()

    plt.title(f'LCOH por Tecnologia de Eletrolisador e Fonte de Energia \n - {name}', fontsize=16, fontweight='bold')

    ax.grid(True, linestyle='--')
    
    ax.set_yticks(np.arange(0, max(max(custo_pem), max(custo_alk), max(custo_soec), max(custo_aem)) + 1.25, 1.25))

    # Função para fechar o gráfico ao pressionar 'q'
    def press(event):
        if event.key == 'q' or 'Q':
            plt.close()

    # Conectar a função de pressionar tecla ao gráfico
    plt.connect('key_press_event', press)
    fig_path = 'C:\\Users\\breno\\Documents\\python_codes\\IC_main\\fig'
    if not os.path.exists(fig_path):
        os.makedirs(fig_path)

    plt.savefig(os.path.join(fig_path, f'LCOH-{name}.png'))
    # Mostrar gráfico

def write_txt(lcoh_pem_pv, lcoh_pem_WindOnshore, lcoh_pem_WindOffshore, lcoh_pem_nuclear,
              lcoh_alk_pv, lcoh_alk_WindOnshore, lcoh_alk_WindOffshore, lcoh_alk_nuclear,
              lcoh_soec_pv, lcoh_soec_WindOnshore, lcoh_soec_WindOffshore, lcoh_soec_nuclear,
              lcoh_aem_pv, lcoh_aem_WindOnshore, lcoh_aem_WindOffshore, lcoh_aem_nuclear, name_file):
    
    lcoh_dict = {
        'Solar PEM': lcoh_pem_pv,
        'WindOnshore PEM': lcoh_pem_WindOnshore,
        'WindOffshore PEM': lcoh_pem_WindOffshore,
        'Nuclear PEM': lcoh_pem_nuclear,
        'Solar AWE': lcoh_alk_pv,
        'WindOnshore AWE': lcoh_alk_WindOnshore,
        'WindOffshore AWE': lcoh_alk_WindOffshore,
        'Nuclear AWE': lcoh_alk_nuclear,
        'Solar SOEC': lcoh_soec_pv,
        'WindOnshore SOEC': lcoh_soec_WindOnshore,
        'WindOffshore SOEC': lcoh_soec_WindOffshore,
        'Nuclear SOEC': lcoh_soec_nuclear,
        'Solar AEM': lcoh_aem_pv,
        'WindOnshore AEM': lcoh_aem_WindOnshore,
        'WindOffshore AEM': lcoh_aem_WindOffshore,
        'Nuclear AEM': lcoh_aem_nuclear
    }

    output_path = 'C:\\Users\\breno\\Documents\\python_codes\\IC_main\\lcoh-txt'  # Caminho da pasta de destino
    file_name = f'LCOH-{name_file}.txt'  # Nome do arquivo
    file_path = os.path.join(output_path, file_name)
    
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    try:
        with open(os.path.join(output_path, file_name), 'w') as arquivo:
            i = 0
            for key, value in lcoh_dict.items():
                i += 1
                arquivo.write(f'{key}: {value:.2f} $/kg\n')
                if i % 4 == 0:
                    arquivo.write('\n')

    except Exception as e:
        print('Errorouuuuuuu: ' + str(e))

def write_to_sheet(sheet_name, txt_file_name, writer):
    file_path = (f'lcoh-txt/LCOH-{txt_file_name}.txt')
    data = {'Technology': [], 'LCOH': []}

    # Lendo o arquivo txt e preenchendo o dicionário
    with open(file_path, 'r') as file:
        for line in file:
            if line.strip():
                parts = line.split(':')
                tech = parts[0].strip()
                lcoh = float(parts[1].split('$')[0].strip())
                data['Technology'].append(tech)
                data['LCOH'].append(lcoh)

    # Convertendo o dicionário para DataFrame
    df = pd.DataFrame(data)
    # Escrevendo o DataFrame no Excel, em uma sheet específica para o ano
    df.to_excel(writer, sheet_name=str(sheet_name), index=False)

    # Abrindo a sheet recém-criada para aplicar a formatação
    #workbook = writer.book
    worksheet = writer.book[sheet_name]

    # Centralizando os dados em todas as células da planilha
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # Obtém a letra da coluna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Fator de ajuste para garantir espaço extra
        worksheet.column_dimensions[column_letter].width = adjusted_width

def write_excel(file_name, txt_file_name, sheet_name):
    #Caminho do arquivo Excel final
    directory = 'C:\\Users\\breno\\Documents\\python_codes\\IC_main\\excel_file'

    if not os.path.exists(directory):
        os.makedirs(directory)

    excel_path = os.path.join(directory, f'{file_name}.xlsx')

    # Criando um ExcelWriter para poder escrever múltiplas sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Chamar a função write_to_sheet para cada ano necessário
        # write_to_sheet first argument is the .txt file name. 
        # Change to pessimist, convservative optmist 
        write_to_sheet(sheet_name, txt_file_name, writer)

def wh2(pot, energy1kg, nome, cf):
    aep = pot*24*365*0.913 # 91.3% Fator de capacidade eletrolisadores
    wh2 = aep/energy1kg
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
    print(f'{nome} WH2: {wh2:.2f} kg')
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
    return wh2

def wh2_no_storage(pot, energy1kg, nome, cf):
    aep = pot*24*365*0.913*cf # 91.3% Fator de capacidade eletrolisadores
    wh2_no_storage = aep/energy1kg
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
    print(f'{nome} WH2: {wh2_no_storage:.2f} kg')
    print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
    return wh2_no_storage

# Calculo do Levelised Cosf Of Hydrogen
def lcoh(tot_capex_el, tot_opex_el, tot_capex_energy, tot_opex_energy, ciclos, wh2, t, nome, energia_nome, wacc):
    """
    Consideracoes:
    Stack replacement costs  = 50% Total CAPEX
    Installation cost = 12% Total CAPEX
    Indirect cost = 20% Total CAPEX
    """
    #capexEletrolisador = tot_capex_el + tot_capex_el*0.5*ciclos + 0.12*tot_capex_el + 0.2*tot_capex_el
    capexEletrolisador = tot_capex_el*(1.32+0.5*ciclos)
    capex_total = capexEletrolisador + tot_capex_energy
    opex_ano = tot_opex_el + tot_opex_energy
    ahp_ano = wh2
    
    print(f'{nome} - {energia_nome} Capex total: {capex_total*10**-6:.2f}')
    print(f'{nome} - {energia_nome} Opex anual: {opex_ano*10**-6:.2f}')
    opex = 0
    total_lcoh = 0
    ahp_total = 0

    # Calcula o opex total considerando o wacc ao longo de t anos
    for i in range(1,t+1):
        op = (opex_ano)/((1 + wacc)**i)
        opex += op 

    # Calcula a produção total de hidrogenio ao longo de t anos
    for i in range(1, t+1):
        ahp_i = ahp_ano/((1 + wacc)**i)
        ahp_total += ahp_i

    total_lcoh = ((capex_total) + (opex))/(ahp_total)

    print(f'A conta é {nome} - {energia_nome} ({capex_total*10**-6:.2f} + {opex*10**-6:.2f})/({ahp_total*10**-3:.2f} t) = LCOH = {total_lcoh:.2f} $/kg')
    print('---------------------------------------------%')
    return total_lcoh